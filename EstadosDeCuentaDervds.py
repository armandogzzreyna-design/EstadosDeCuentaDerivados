import streamlit as st
import pandas as pd
import io
from datetime import date, timedelta
import os

# ============================================================================
# CONFIGURACIÓN INICIAL
# ============================================================================

st.set_page_config(
    page_title="Validador de Estados de Cuenta - Derivados",
    layout="wide",
    initial_sidebar_state="collapsed"
)

import re

try:
    import pdfplumber
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError as e:
    st.error(
        "Falta instalar una dependencia de Python. "
        "Ejecuta: pip install -r requirements.txt"
    )
    st.caption(f"Detalle: {e}")
    st.stop()

# ============================================================================
# ESTILOS Y CONSTANTES
# ============================================================================

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
}

SIEFORE_A_SB = {
    "60": "SB60", "60-64": "SB60", "6064": "SB60",
    "65": "SB65", "65-69": "SB65", "6569": "SB65",
    "70": "SB70", "70-74": "SB70", "7074": "SB70",
    "75": "SB75", "75-79": "SB75", "7579": "SB75",
    "80": "SB80", "80-84": "SB80", "8084": "SB80",
    "85": "SB85", "85-89": "SB85", "8589": "SB85",
    "90": "SB90", "90-94": "SB90", "9094": "SB90",
    "95": "SB95", "95-99": "SB95", "9599": "SB95",
    "INICIAL": "SBIN", "INCIAL": "SBIN", "IN": "SBIN",
    "INICIA": "SBIN", "INCIAL": "SBIN", "IN": "SBIN",
}

TOLERANCIA = 0.01

# Tabla de referencia de cuentas (REF_DATA del script original)
REF_DATA = [
    ("SCOTIA","SIEFORE INVERCAP 95","1008375","BCSCOT95F"),
    ("SCOTIA","SIEFORE INVERCAP 60","1008793","BCSCOT60F"),
    ("SCOTIA","SIEFORE INVERCAP 65","10081002","BCSCOT65F"),
    ("SCOTIA","SIEFORE INVERCAP 70","10024618","BCSCOT70F"),
    ("SCOTIA","SIEFORE INVERCAP 75","1001717","BCSCOT75F"),
    ("SCOTIA","SIEFORE INVERCAP 80","10086617","BCSCOT80F"),
    ("SCOTIA","SIEFORE INVERCAP 85","1003533","BCSCOT85F"),
    ("SCOTIA","SIEFORE INVERCAP 90","10091375","BCSCOT90F"),
    ("SCOTIA","SIEFORE INVERCAP IN","10096697","BCSCOTINF"),
    ("SCOTIA","SIEFORE INVERCAP 60","8793","BCSCOT60"),
    ("SCOTIA","SIEFORE INVERCAP 65","81002","BCSCOT65"),
    ("SCOTIA","SIEFORE INVERCAP 70","24618","BCSCOT70"),
    ("SCOTIA","SIEFORE INVERCAP 75","1717","BCSCOT75"),
    ("SCOTIA","SIEFORE INVERCAP 80","86617","BCSCOT80"),
    ("BCSANT","SIEFORE INVERCAP 95","08375SIC1C","BCSANT95F"),
    ("BCSANT","SIEFORE INVERCAP 60","08793SIC2C","BCSANT60F"),
    ("BCSANT","SIEFORE INVERCAP 65","81002SIC6C","BCSANT65F"),
    ("BCSANT","SIEFORE INVERCAP 70","24618SIC7C","BCSANT70F"),
    ("BCSANT","SIEFORE INVERCAP 75","01717SIC3C","BCSANT75F"),
    ("BCSANT","SIEFORE INVERCAP 80","86617SIC8C","BCSANT80F"),
    ("BCSANT","SIEFORE INVERCAP 85","03533SIC4C","BCSANT85F"),
    ("BCSANT","SIEFORE INVERCAP 90","91375SIC9C","BCSANT90F"),
    ("BCSANT","SIEFORE INVERCAP IN","96697SICBC","BCSANTINF"),
    ("BCSANT","SIEFORE INVERCAP 65","810022IC6I","BCSANT65"),
    ("BCSANT","SIEFORE INVERCAP 60","087932IC2I","BCSANT60"),
    ("BCSANT","SIEFORE INVERCAP 70","246182IC7I","BCSANT70"),
    ("BCSANT","SIEFORE INVERCAP 75","017172IC3I","BCSANT75"),
    ("BCSANT","SIEFORE INVERCAP 80","866172IC8I","BCSANT80"),
    ("BCSANT","SIEFORE INVERCAP 85","035332IC4I","BCSANT85"),
    ("BCSANT","SIEFORE INVERCAP 90","913752IC9I","BCSANT90"),
    ("BCSANT","SIEFORE INVERCAP IN","966972ICBI","BCSANTIN"),
    ("BCSANT","SIEFORE INVERCAP 95","083752IC1I","BCSANT95"),
    ("SANTANDER CME","SIEFORE INVERCAP BP","UKN99","BSANCME10"),
    ("SANTANDER CME","SIEFORE INVERCAP 95","UKJ99","BSANCME195"),
    ("SANTANDER CME","SIEFORE INVERCAP 60","UKK99","BSANCME160"),
    ("SANTANDER CME","SIEFORE INVERCAP 65","VLM99","BSANCME165"),
    ("SANTANDER CME","SIEFORE INVERCAP 70","VLN99","BSANCME170"),
    ("SANTANDER CME","SIEFORE INVERCAP 75","UKL99","BSANCME175"),
    ("SANTANDER CME","SIEFORE INVERCAP 80","VLO99","BSANCME180"),
    ("SANTANDER CME","SIEFORE INVERCAP 85","UKM99","BSANCME185"),
    ("SANTANDER CME","SIEFORE INVERCAP 90","VLP99","BSANCME190"),
    ("SANTANDER CME","SIEFORE INVERCAP IN","VLQ99","BSANCME1IN"),
    ("GOLDMAN","SIEFORE INVERCAP BP","191086",""),
    ("GOLDMAN","SIEFORE INVERCAP 95","191089","GOLDMAN95"),
    ("GOLDMAN","SIEFORE INVERCAP 60","191091","GOLDMAN60"),
    ("GOLDMAN","SIEFORE INVERCAP 65","105729","GOLDMAN65"),
    ("GOLDMAN","SIEFORE INVERCAP 70","105732","GOLDMAN70"),
    ("GOLDMAN","SIEFORE INVERCAP 75","191087","GOLDMAN75"),
    ("GOLDMAN","SIEFORE INVERCAP 80","105733","GOLDMAN80"),
    ("GOLDMAN","SIEFORE INVERCAP 85","191090","GOLDMAN85"),
    ("GOLDMAN","SIEFORE INVERCAP 90","105735","GOLDMAN90"),
    ("GOLDMAN","SIEFORE INVERCAP IN","105736","GOLDMANIN"),
]

df_ref = pd.DataFrame(REF_DATA, columns=["Counterparty","Portfolio","Account","Portafolio"])

# ============================================================================
# FUNCIONES UTILITARIAS
# ============================================================================

def archivos_subidos_a_dict(files):
    """Convierte archivos de Streamlit a bytes sin consumir el stream."""
    return {f.name: f.getvalue() for f in files} if files else {}

@st.cache_data(show_spinner=False)
def leer_excel_cacheado(file_bytes):
    """Lee Excel desde bytes y evita releerlo en cada refresco de Streamlit."""
    return pd.read_excel(io.BytesIO(file_bytes))

def normalizar_texto(valor):
    """Normaliza texto para comparar columnas y llaves con mayor tolerancia."""
    return re.sub(r"\s+", " ", str(valor).strip().upper())

def numero_o_none(valor):
    """Convierte un valor a número; si no se puede, regresa None."""
    if pd.isna(valor):
        return None
    try:
        return float(valor)
    except (TypeError, ValueError):
        return None

def preparar_aims_ref(df_aims_ref):
    """Normaliza el archivo AIMS una sola vez para acelerar la validación."""
    if df_aims_ref is None or df_aims_ref.empty:
        return None
    
    df = df_aims_ref.copy()
    df.columns = [str(c).strip() for c in df.columns]
    
    requeridas = {"SOCIO", "SIEFORE", "Contribución AIM Subcuenta", "AIM Cuenta"}
    if not requeridas.issubset(set(df.columns)):
        return df
    
    df["SOCIO_KEY"] = df["SOCIO"].map(normalizar_texto)
    df["SIEFORE_KEY"] = df["SIEFORE"].map(normalizar_texto)
    df["AIM_SUBCUENTA_NUM"] = pd.to_numeric(df["Contribución AIM Subcuenta"], errors="coerce")
    df["AIM_CUENTA_NUM"] = pd.to_numeric(df["AIM Cuenta"], errors="coerce")
    return df

def preparar_colateral_ref(df_colat_ref):
    """Normaliza IA Colateral y arma llaves de búsqueda por socio/SIEFORE."""
    if df_colat_ref is None or df_colat_ref.empty:
        return None
    
    df = df_colat_ref.copy()
    primera_col = df.iloc[:, 0].astype(str).map(normalizar_texto)
    df["_CODIGO"] = primera_col
    df["_SIEFORE"] = primera_col.map(codigo_a_sb)
    df["_SOCIO"] = ""
    df.loc[primera_col.str.contains("BCSCOT", na=False), "_SOCIO"] = "SCOTIA"
    df.loc[primera_col.str.contains("BCSANT", na=False), "_SOCIO"] = "SANTANDER"
    
    if df.shape[1] > 16:
        df["_MONTO_COLATERAL"] = pd.to_numeric(df.iloc[:, 16], errors="coerce")
    else:
        df["_MONTO_COLATERAL"] = pd.NA
    
    return df[(df["_SIEFORE"].notna()) & (df["_SOCIO"] != "")].copy()

def limpiar_numero(texto):
    """Convierte cadenas numéricas a float, soporta múltiples formatos."""
    t = str(texto).strip()
    negativo = t.startswith("(") and t.endswith(")")
    t = t.strip("()")
    t = re.sub(r'(?<=\d) (?=\d)', '', t)
    if re.search(r",\d{1,2}$", t):
        t = t.replace(".", "").replace(",", ".")
    else:
        t = t.replace(",", "")
    try:
        valor = float(t)
    except ValueError:
        valor = 0.0
    return -valor if negativo else valor

@st.cache_data(show_spinner=False)
def texto_paginas(pdf_bytes):
    """Extrae texto de PDF desde bytes."""
    paginas = []
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                paginas.append(page.extract_text() or "")
    except Exception as e:
        return []
    return paginas

def normalizar_account(acc):
    """Normaliza números de cuenta."""
    return re.sub(r'^0+', '', str(acc).strip().upper())

def extraer_sb(siefore_raw):
    """Extrae SB60, SB65, etc. del texto de siefore."""
    s = siefore_raw.upper()
    s = re.sub(r'SIEFORE INVERCAP BASICA', '', s)
    s = s.replace(",", "").strip()
    for key, val in SIEFORE_A_SB.items():
        if s == key or s.startswith(key + " ") or s.startswith(key + "-"):
            return val
    primera = s.split()[0] if s.split() else s
    return SIEFORE_A_SB.get(primera, s.strip())

def codigo_a_sb(codigo):
    """Convierte BCSCOT95F -> SB95, BCSANT75F -> SB75, BCSCOTINF -> SBIN"""
    codigo = str(codigo).upper().strip()
    codigo = codigo.replace(".0", "")
    match = re.search(r'(\d+F|INF)$', codigo)
    if match:
        sufijo = match.group(1)
        if sufijo == "INF":
            return "SBIN"
        sufijo_num = sufijo.replace("F", "")
        return f"SB{sufijo_num}"
    return None

# ============================================================================
# PARSERS PDF (LOGICA ORIGINAL)
# ============================================================================

@st.cache_data(show_spinner=False)
def parse_goldman_cme(pdf_files_dict):
    """Parser para Goldman CME."""
    resultados = []
    if not pdf_files_dict:
        return resultados
    
    for filename, pdf_bytes in pdf_files_dict.items():
        paginas = texto_paginas(pdf_bytes)
        for texto in paginas:
            if not re.search(r'JOURNAL ENTRIES', texto, re.IGNORECASE):
                continue
            if re.search(r'NO ACTIVITY', texto, re.IGNORECASE) and \
               not re.search(r'FUNDS\s+(PAID|RECEIVED)', texto, re.IGNORECASE):
                continue
            
            m_sief = re.search(r'SIEFORE INVERCAP BASICA\s+([\w][\w\-]*)', texto, re.IGNORECASE)
            siefore_raw = m_sief.group(0).strip().upper() if m_sief else ""
            
            m_acc = re.search(r'ACCOUNT NUMBER[:\s]+([\d\s]+)', texto, re.IGNORECASE)
            account_raw = re.sub(r'\s+', '', m_acc.group(1)).split("\n")[0].strip() if m_acc else ""
            
            es_paid = bool(re.search(r'FUNDS\s+PAID', texto, re.IGNORECASE))
            es_received = bool(re.search(r'FUNDS\s+RECEIVED', texto, re.IGNORECASE))
            
            if not es_paid and not es_received:
                continue
            
            m_monto = re.search(r'USD\s+([\(]?[\d,\.]+[\)]?)', texto)
            if not m_monto:
                continue
            
            monto = abs(limpiar_numero(m_monto.group(1)))
            
            if es_received:
                monto = -monto
            
            if monto == 0 or not siefore_raw:
                continue
            
            if "PROCESSING" in texto.upper():
                continue
            
            resultados.append({
                "Counterparty": "GOLDMAN",
                "siefore_raw": siefore_raw,
                "account_raw": account_raw,
                "Currency": "USD",
                "Monto": monto,
            })
    
    return resultados

@st.cache_data(show_spinner=False)
def parse_santander_cme(pdf_files_dict):
    """Parser para Santander CME."""
    resultados = []
    if not pdf_files_dict:
        return resultados
    
    for filename, pdf_bytes in pdf_files_dict.items():
        paginas = texto_paginas(pdf_bytes)
        for texto in paginas:
            if not re.search(r'Saldo Libre', texto, re.IGNORECASE):
                continue
            
            m_cuenta = re.search(r'CUENTA:\s*([A-Z0-9]+)', texto)
            account = m_cuenta.group(1).strip() if m_cuenta else ""
            
            m_descr = re.search(r'DESCR\.\s*CTA:\s*(SIEFORE INVERCAP[^\n]+)', texto, re.IGNORECASE)
            siefore_raw = m_descr.group(1).strip().upper() if m_descr else ""
            siefore_raw = re.sub(r'\s*SA DE CV.*$', '', siefore_raw).strip()
            
            m_saldo = re.search(r'Saldo Libre disposici[o\xf3n]+\s+([\-\d\.\,]+)', texto, re.IGNORECASE)
            if not m_saldo:
                continue
            
            monto = limpiar_numero(m_saldo.group(1))
            
            if monto == 0 or not siefore_raw:
                continue
            
            resultados.append({
                "Counterparty": "SANTANDER CME",
                "siefore_raw": siefore_raw,
                "account_raw": account,
                "Currency": "USD",
                "Monto": monto,
            })
    
    visto = {}
    for r in resultados:
        visto[r["account_raw"]] = r
    resultados = list(visto.values())
    
    return resultados

@st.cache_data(show_spinner=False)
def parse_santander_mexder(pdf_files_dict):
    """Parser para Santander MEXDER."""
    resultados = []
    if not pdf_files_dict:
        return resultados
    
    for filename, pdf_bytes in pdf_files_dict.items():
        paginas = texto_paginas(pdf_bytes)
        texto_completo = "\n".join(paginas)
        
        m_sief = re.search(r'(SIEFORE INVERCAP BASICA[\s\w]+?)(?:\s+SA DE CV|\s+RFC|\s*\n)', texto_completo, re.IGNORECASE)
        siefore_raw = m_sief.group(1).strip().upper() if m_sief else ""
        
        m_acc = re.search(r'ACCOUNT NO\.?:\s*([A-Z0-9]+)', texto_completo, re.IGNORECASE)
        account = m_acc.group(1).strip() if m_acc else ""
        
        m_valor = re.search(r"VALOR TOTAL DE LA CUENTA[^\n]*?([\-]?[\d,\.]+)\s*$", texto_completo, re.IGNORECASE | re.MULTILINE)
        if not m_valor:
            continue
        
        monto = limpiar_numero(m_valor.group(1))
        
        if monto == 0 or not siefore_raw:
            continue
        
        resultados.append({
            "Counterparty": "SANTANDER MEXDER",
            "siefore_raw": siefore_raw,
            "account_raw": account,
            "Currency": "MXN",
            "Monto": monto,
        })
    
    return resultados

@st.cache_data(show_spinner=False)
def parse_scotia_mexder(pdf_files_dict):
    """Parser para Scotia MEXDER."""
    resultados = []
    if not pdf_files_dict:
        return resultados
    
    for filename, pdf_bytes in pdf_files_dict.items():
        paginas = texto_paginas(pdf_bytes)
        texto_completo = "\n".join(paginas)
        
        m_sief = re.search(r'SIEFORE INVERCAP BASICA,?\s+[\w][\w\-\s]*', texto_completo, re.IGNORECASE)
        siefore_raw = m_sief.group(0).strip().upper() if m_sief else ""
        siefore_raw = siefore_raw.replace(",", "").strip()
        siefore_raw = re.sub(r'\s+(AV|COL|RFC|CP|SA|DEL|LORENZO|CUIDAD|ACCOUNT|DATE|PAGE)\b.*', '', siefore_raw, flags=re.IGNORECASE).strip()
        
        m_acc = re.search(r'ACCOUNT NUMBER[:\s]+([\d]+)', texto_completo, re.IGNORECASE)
        account = m_acc.group(1).strip() if m_acc else ""
        
        m_margin = re.search(r'MARGIN\s+DEFAULT/EXCESS\s+([\d,\.\s]+?)\s+(CR|DR)', texto_completo, re.IGNORECASE)
        if not m_margin:
            continue
        
        monto = limpiar_numero(m_margin.group(1))
        crdr = m_margin.group(2).upper()
        
        monto = abs(monto)
        if crdr == "DR":
            monto = -monto
        
        if monto == 0 or not siefore_raw:
            continue
        
        resultados.append({
            "Counterparty": "SCOTIA",
            "siefore_raw": siefore_raw,
            "account_raw": account,
            "Currency": "MXN",
            "Monto": monto,
        })
    
    return resultados

# ============================================================================
# ENRIQUECIMIENTO DE DATOS
# ============================================================================

def enriquecer(registros, df_ref, fecha_val):
    """Enriquece registros con datos de la tabla de referencia."""
    filas = []
    fecha_str = fecha_val.strftime("%d/%m/%Y")
    
    for r in registros:
        cp = r["Counterparty"]
        acc_raw = r["account_raw"]
        
        CP_ALIAS = {
            "SANTANDER MEXDER": "BCSANT",
            "GOLDMAN CME": "GOLDMAN",
            "SANTANDER CME": "SANTANDER CME",
            "SCOTIA": "SCOTIA",
        }
        cp_ref = CP_ALIAS.get(cp.upper(), cp.upper())
        sub = df_ref[df_ref["Counterparty"].str.upper() == cp_ref.upper()]
        
        match = sub[sub["Account"].str.upper() == acc_raw.upper()]
        
        if match.empty:
            match = sub[sub["Account"].apply(normalizar_account) == normalizar_account(acc_raw)]
        
        if not match.empty:
            ref = match.iloc[0]
            portfolio = ref["Portfolio"]
            account = ref["Account"]
            portafolio = ref["Portafolio"]
        else:
            portfolio = r["siefore_raw"]
            account = acc_raw
            portafolio = ""
        
        filas.append({
            "Counterparty": cp,
            "Portfolio": portfolio,
            "Date": fecha_str,
            "Currency": r["Currency"],
            "Account": account,
            "Monto": r["Monto"],
            "Portafolio": portafolio,
        })
    
    return pd.DataFrame(filas, columns=["Counterparty","Portfolio","Date","Currency","Account","Monto","Portafolio"])

# ============================================================================
# VALIDACIONES
# ============================================================================

@st.cache_data(show_spinner=False)
def procesar_aims_pdf(pdf_files_dict_scotia, pdf_files_dict_santander):
    """Extrae montos AIMS de los PDFs."""
    aims_pdf = []
    
    # Scotia MEXDER: RISK MAINTENANCE
    for filename, pdf_bytes in (pdf_files_dict_scotia or {}).items():
        paginas = texto_paginas(pdf_bytes)
        texto_completo = "\n".join(paginas)
        m_sief = re.search(r'SIEFORE INVERCAP BASICA,?\s+[\w][\w\-\s]*', texto_completo, re.IGNORECASE)
        if not m_sief:
            continue
        siefore_raw = m_sief.group(0).strip().upper().replace(",", "")
        siefore_raw = siefore_raw.split("\n")[0].strip()
        siefore_raw = re.sub(r'\s+(AV|COL|RFC|CP|SA|DEL|LORENZO|CUIDAD|ACCOUNT|DATE|PAGE)\b.*', '', siefore_raw, flags=re.IGNORECASE).strip()
        sb = extraer_sb(siefore_raw)
        m_risk = re.search(r'RISK\s+MAINTENANCE\s+([\d,\.\s]+?)\s+(CR|DR)', texto_completo, re.IGNORECASE)
        if not m_risk:
            continue
        monto = abs(limpiar_numero(m_risk.group(1)))
        if monto == 0:
            continue
        aims_pdf.append({"SIEFORE": sb, "SOCIO": "SCOTIA", "Monto_edo": monto})
    
    # Santander MEXDER: MARGEN CAMARA T + 1
    for filename, pdf_bytes in (pdf_files_dict_santander or {}).items():
        paginas = texto_paginas(pdf_bytes)
        texto_completo = "\n".join(paginas)
        m_sief = re.search(r'(SIEFORE INVERCAP BASICA[\s\w]+?)(?:\s+SA DE CV|\s+RFC|\s*\n)', texto_completo, re.IGNORECASE)
        siefore_raw = m_sief.group(1).strip().upper() if m_sief else ""
        if not siefore_raw:
            continue
        sb = extraer_sb(siefore_raw)
        m_margen = re.search(r'MARGEN\s+CAMARA\s+T\s*\+\s*1[^\n]*?([\d,\.]+)\s*$', texto_completo, re.IGNORECASE | re.MULTILINE)
        if not m_margen:
            continue
        monto = abs(limpiar_numero(m_margen.group(1)))
        if monto == 0:
            continue
        aims_pdf.append({"SIEFORE": sb, "SOCIO": "SANTANDER", "Monto_edo": monto})
    
    return aims_pdf

@st.cache_data(show_spinner=False)
def procesar_colateral_pdf(pdf_files_dict_scotia, pdf_files_dict_santander):
    """Extrae montos de colateral de los PDFs."""
    colateral_pdf = []
    
    # Scotia MEXDER: SECURITIES ON DEPOSIT
    for filename, pdf_bytes in (pdf_files_dict_scotia or {}).items():
        paginas = texto_paginas(pdf_bytes)
        texto_completo = "\n".join(paginas)
        m_sief = re.search(r'SIEFORE INVERCAP BASICA,?\s+[\w][\w\-\s]*', texto_completo, re.IGNORECASE)
        if not m_sief:
            continue
        siefore_raw = m_sief.group(0).strip().upper().replace(",", "")
        siefore_raw = siefore_raw.split("\n")[0].strip()
        siefore_raw = re.sub(r'\s+(AV|COL|RFC|CP|SA|DEL|LORENZO|CUIDAD|ACCOUNT|DATE|PAGE)\b.*', '', siefore_raw, flags=re.IGNORECASE).strip()
        sb = extraer_sb(siefore_raw)
        m_sec = re.search(r'SECURITIES\s+ON\s+DEPOSIT\s+([\d,\.\s]+?)\s+(CR|DR)', texto_completo, re.IGNORECASE)
        if not m_sec:
            continue
        monto = abs(limpiar_numero(m_sec.group(1)))
        if monto == 0:
            continue
        colateral_pdf.append({"SIEFORE": sb, "SOCIO": "SCOTIA", "Monto_edo": monto})
    
    # Santander MEXDER: VALOR COLATERALES EN MARGEN CAMARA
    for filename, pdf_bytes in (pdf_files_dict_santander or {}).items():
        paginas = texto_paginas(pdf_bytes)
        texto_completo = "\n".join(paginas)
        m_sief = re.search(r'(SIEFORE INVERCAP BASICA[\s\w]+?)(?:\s+SA DE CV|\s+RFC|\s*\n)', texto_completo, re.IGNORECASE)
        siefore_raw = m_sief.group(1).strip().upper() if m_sief else ""
        if not siefore_raw:
            continue
        sb = extraer_sb(siefore_raw)
        m_colat = re.search(r'VALOR COLATERALES EN MARGEN CAMARA[^\n]*?([\d,\.]+)\s*$', texto_completo, re.IGNORECASE | re.MULTILINE)
        if not m_colat:
            continue
        monto = abs(limpiar_numero(m_colat.group(1)))
        if monto == 0:
            continue
        colateral_pdf.append({"SIEFORE": sb, "SOCIO": "SANTANDER", "Monto_edo": monto})
    
    return colateral_pdf

def validar_aims(aims_pdf, df_aims_ref):
    """Valida montos AIMS contra archivo de referencia."""
    filas_aims = []
    indices_usados = set()
    df_aims_ref = preparar_aims_ref(df_aims_ref)
    
    for item in aims_pdf:
        sb = normalizar_texto(item["SIEFORE"])
        socio = normalizar_texto(item["SOCIO"])
        monto_edo = item["Monto_edo"]
        aim_subcuenta = None
        aim_cuenta = None
        
        if df_aims_ref is not None and {"SOCIO_KEY", "SIEFORE_KEY"}.issubset(df_aims_ref.columns):
            candidatos = df_aims_ref[
                (df_aims_ref["SOCIO_KEY"] == socio) &
                (df_aims_ref["SIEFORE_KEY"] == sb)
            ]
            candidatos = candidatos[~candidatos.index.isin(indices_usados)]
            if not candidatos.empty:
                idx = candidatos.index[0]
                indices_usados.add(idx)
                aim_subcuenta = numero_o_none(candidatos.loc[idx, "AIM_SUBCUENTA_NUM"])
                aim_cuenta = numero_o_none(candidatos.loc[idx, "AIM_CUENTA_NUM"])
        elif df_aims_ref is not None:
            st.warning("El archivo AIMS no tiene las columnas esperadas para validar.")
        
        dif_sub = round(monto_edo - aim_subcuenta, 2) if aim_subcuenta is not None else None
        dif_cuenta = round(monto_edo - aim_cuenta, 2) if aim_cuenta is not None else None
        
        if dif_sub is not None and dif_cuenta is not None:
            estatus = "OK" if abs(dif_sub) <= TOLERANCIA and abs(dif_cuenta) <= TOLERANCIA else "DIFERENCIA"
        else:
            estatus = ""
        
        filas_aims.append({
            "SIEFORE": sb,
            "SOCIO": socio,
            "Monto Estado": monto_edo,
            "AIM Subcuenta": aim_subcuenta if aim_subcuenta is not None else "Sin posición",
            "AIM Cuenta": aim_cuenta if aim_cuenta is not None else "Sin posición",
            "Diferencia Subcuenta": dif_sub,
            "Diferencia Cuenta": dif_cuenta,
            "Estatus": estatus,
        })
    
    df_aims_val = pd.DataFrame(filas_aims, columns=[
        "SIEFORE", "SOCIO", "Monto Estado",
        "AIM Subcuenta", "AIM Cuenta",
        "Diferencia Subcuenta", "Diferencia Cuenta", "Estatus"
    ])
    
    if not df_aims_val.empty:
        df_aims_val = df_aims_val.sort_values(["SOCIO", "SIEFORE"]).reset_index(drop=True)
    
    return df_aims_val

def validar_colateral(colateral_pdf, df_colat_ref):
    """Valida montos de colateral."""
    filas_colat = []
    indices_usados_colat = set()
    df_colat_ref = preparar_colateral_ref(df_colat_ref)
    
    for item in colateral_pdf:
        sb = normalizar_texto(item["SIEFORE"])
        socio = normalizar_texto(item["SOCIO"])
        monto_edo = item["Monto_edo"]
        monto_val = None
        
        if df_colat_ref is not None and not df_colat_ref.empty:
            candidatos = df_colat_ref[
                (df_colat_ref["_SOCIO"] == socio) &
                (df_colat_ref["_SIEFORE"] == sb)
            ]
            candidatos = candidatos[~candidatos.index.isin(indices_usados_colat)]
            if not candidatos.empty:
                idx = candidatos.index[0]
                indices_usados_colat.add(idx)
                monto_val = numero_o_none(candidatos.loc[idx, "_MONTO_COLATERAL"])
        
        if monto_edo == 0 and monto_val is None:
            continue
        
        dif = round(monto_edo - monto_val, 2) if monto_val is not None else None
        
        if dif is not None:
            estatus = "OK" if abs(dif) <= TOLERANCIA else "DIFERENCIA"
        else:
            estatus = ""
        
        filas_colat.append({
            "SIEFORE": sb,
            "SOCIO": socio,
            "Monto Estado": monto_edo,
            "IA Colateral": monto_val if monto_val is not None else "Sin datos",
            "Diferencia": dif,
            "Estatus": estatus,
        })
    
    df_colat_val = pd.DataFrame(filas_colat, columns=[
        "SIEFORE", "SOCIO", "Monto Estado",
        "IA Colateral", "Diferencia", "Estatus"
    ])
    
    if not df_colat_val.empty:
        df_colat_val = df_colat_val.sort_values(["SOCIO", "SIEFORE"]).reset_index(drop=True)
    
    return df_colat_val

# ============================================================================
# GENERACIÓN DE EXCEL
# ============================================================================

def generar_excel(df_resultado, df_aims_val, df_colat_val, fecha_val):
    """Genera archivo Excel con formato profesional."""
    hoy = date.today()
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Llamadas Retiros"
    
    # Estilos
    FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
    FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)
    FONT_ROJO = Font(color="C00000", bold=True)
    FONT_VERDE = Font(color="375623", bold=True)
    
    # --- HOJA 1: Llamadas Retiros ---
    COLUMNAS = ["Counterparty","Portfolio","Date","Currency","Account","Monto","Portafolio"]
    ANCHOS = [20, 30, 14, 10, 18, 16, 16]
    
    for col, (nombre_col, ancho) in enumerate(zip(COLUMNAS, ANCHOS), 1):
        celda = ws.cell(row=1, column=col, value=nombre_col)
        celda.fill = FILL_HEADER
        celda.font = FONT_HEADER
        celda.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = ancho
    
    ws.row_dimensions[1].height = 20
    
    for fila_idx, row in df_resultado.iterrows():
        excel_row = fila_idx + 2
        for col_idx, col_nombre in enumerate(COLUMNAS, 1):
            valor = row[col_nombre]
            celda = ws.cell(row=excel_row, column=col_idx, value=valor)
            celda.alignment = Alignment(vertical="center")
            if col_nombre == "Monto" and isinstance(valor, (int, float)):
                celda.number_format = '#,##0.00'
                celda.font = FONT_ROJO if valor < 0 else FONT_VERDE
    
    fila_pie = len(df_resultado) + 3
    celda_pie = ws.cell(
        row=fila_pie, column=1,
        value=(
            f"Generado: {hoy.strftime('%d/%m/%Y')}  |  "
            f"Valuacion: {fecha_val.strftime('%d/%m/%Y')}  |  "
            f"Registros: {len(df_resultado)}"
        )
    )
    celda_pie.font = Font(italic=True, color="808080", size=9)
    ws.merge_cells(start_row=fila_pie, start_column=1, end_row=fila_pie, end_column=7)
    
    # --- HOJA 2: Validación AIMS ---
    ws2 = wb.create_sheet(title="Validacion AIMS")
    COLS_AIMS = ["SIEFORE", "SOCIO", "Monto Estado", "AIM Subcuenta", "AIM Cuenta", "Diferencia Subcuenta", "Diferencia Cuenta", "Estatus"]
    ANCHOS_AIMS = [10, 14, 20, 20, 20, 22, 20, 14]
    COLS_NUM = {"Monto Estado", "AIM Subcuenta", "AIM Cuenta", "Diferencia Subcuenta", "Diferencia Cuenta"}
    
    for col, (nombre_col, ancho) in enumerate(zip(COLS_AIMS, ANCHOS_AIMS), 1):
        celda = ws2.cell(row=1, column=col, value=nombre_col)
        celda.fill = FILL_HEADER
        celda.font = FONT_HEADER
        celda.alignment = Alignment(horizontal="center", vertical="center")
        ws2.column_dimensions[get_column_letter(col)].width = ancho
    
    ws2.row_dimensions[1].height = 20
    
    for fila_idx, row in df_aims_val.iterrows():
        excel_row = fila_idx + 2
        for col_idx, col_nombre in enumerate(COLS_AIMS, 1):
            valor = row[col_nombre]
            if not isinstance(valor, str) and pd.isna(valor):
                valor = None
            celda = ws2.cell(row=excel_row, column=col_idx, value=valor)
            celda.alignment = Alignment(vertical="center")
            if col_nombre in COLS_NUM and isinstance(valor, (int, float)):
                celda.number_format = '#,##0.00'
                if col_nombre in ("Diferencia Subcuenta", "Diferencia Cuenta"):
                    celda.font = FONT_VERDE if abs(valor) <= TOLERANCIA else FONT_ROJO
            if col_nombre == "Estatus" and isinstance(valor, str) and valor:
                celda.font = FONT_VERDE if valor == "OK" else FONT_ROJO
    
    fila_pie2 = len(df_aims_val) + 3
    celda_pie2 = ws2.cell(
        row=fila_pie2, column=1,
        value=(
            f"Generado: {hoy.strftime('%d/%m/%Y')}  |  "
            f"Valuacion: {fecha_val.strftime('%d/%m/%Y')}"
        )
    )
    celda_pie2.font = Font(italic=True, color="808080", size=9)
    ws2.merge_cells(start_row=fila_pie2, start_column=1, end_row=fila_pie2, end_column=8)
    
    # --- HOJA 3: IA Colateral ---
    ws3 = wb.create_sheet(title="IA Colateral")
    COLS_COLAT = ["SIEFORE", "SOCIO", "Monto Estado", "IA Colateral", "Diferencia", "Estatus"]
    ANCHOS_COLAT = [10, 14, 20, 20, 18, 14]
    COLS_NUM_COL = {"Monto Estado", "IA Colateral", "Diferencia"}
    
    for col, (nombre_col, ancho) in enumerate(zip(COLS_COLAT, ANCHOS_COLAT), 1):
        celda = ws3.cell(row=1, column=col, value=nombre_col)
        celda.fill = FILL_HEADER
        celda.font = FONT_HEADER
        celda.alignment = Alignment(horizontal="center", vertical="center")
        ws3.column_dimensions[get_column_letter(col)].width = ancho
    
    ws3.row_dimensions[1].height = 20
    
    for fila_idx, row in df_colat_val.iterrows():
        excel_row = fila_idx + 2
        for col_idx, col_nombre in enumerate(COLS_COLAT, 1):
            valor = row[col_nombre]
            if not isinstance(valor, str) and pd.isna(valor):
                valor = None
            celda = ws3.cell(row=excel_row, column=col_idx, value=valor)
            celda.alignment = Alignment(vertical="center")
            if col_nombre in COLS_NUM_COL and isinstance(valor, (int, float)):
                celda.number_format = '#,##0.00'
                if col_nombre == "Diferencia":
                    celda.font = FONT_VERDE if abs(valor) <= TOLERANCIA else FONT_ROJO
            if col_nombre == "Estatus" and isinstance(valor, str) and valor:
                celda.font = FONT_VERDE if valor == "OK" else FONT_ROJO
    
    fila_pie3 = len(df_colat_val) + 3
    celda_pie3 = ws3.cell(
        row=fila_pie3, column=1,
        value=(
            f"Generado: {hoy.strftime('%d/%m/%Y')}  |  "
            f"Valuacion: {fecha_val.strftime('%d/%m/%Y')}"
        )
    )
    celda_pie3.font = Font(italic=True, color="808080", size=9)
    ws3.merge_cells(start_row=fila_pie3, start_column=1, end_row=fila_pie3, end_column=6)
    
    # Guardar a bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# ============================================================================
# INTERFAZ STREAMLIT
# ============================================================================

def main():
    # Header
    st.markdown("# Validador de Estados de Cuenta - Derivados")
    
    col1, col2 = st.columns([2, 1])
    with col1:
        st.markdown("**Procesamiento y validación de derivados contra AIMS e IA Colateral**")
    with col2:
        fecha_valuacion = st.date_input(
            "Fecha de valuación",
            value=date.today(),
            key="fecha_val"
        )
    
    st.divider()
    
    # ========== ETAPA 1: CARGA DE ARCHIVOS ==========
    st.markdown("## Etapa 1: Carga de Archivos")
    
    tab1, tab2, tab3 = st.tabs(["Estados de Cuenta (PDFs)", "Archivo AIMS", "Archivo IA Colateral"])
    
    # --- TAB 1: PDFS ---
    with tab1:
        st.markdown("### Carga de Estados de Cuenta por Contraparte")
        
        col_gs, col_scme, col_smex, col_sco = st.columns(4)
        
        pdf_goldman = None
        pdf_santander_cme = None
        pdf_santander_mexder = None
        pdf_scotia = None
        
        with col_gs:
            st.markdown("**Goldman CME**")
            files_gs = st.file_uploader("Goldman CME", type="pdf", accept_multiple_files=True, key="goldman")
            pdf_goldman = archivos_subidos_a_dict(files_gs)
        
        with col_scme:
            st.markdown("**Santander CME**")
            files_scme = st.file_uploader("Santander CME", type="pdf", accept_multiple_files=True, key="sant_cme")
            pdf_santander_cme = archivos_subidos_a_dict(files_scme)
        
        with col_smex:
            st.markdown("**Santander MEXDER**")
            files_smex = st.file_uploader("Santander MEXDER", type="pdf", accept_multiple_files=True, key="sant_mex")
            pdf_santander_mexder = archivos_subidos_a_dict(files_smex)
        
        with col_sco:
            st.markdown("**Scotia MEXDER**")
            files_sco = st.file_uploader("Scotia MEXDER", type="pdf", accept_multiple_files=True, key="scotia")
            pdf_scotia = archivos_subidos_a_dict(files_sco)
        
        # Resumen de carga
        st.markdown("### Resumen de Carga")
        col_res1, col_res2, col_res3, col_res4, col_res5 = st.columns(5)
        
        with col_res1:
            st.metric("Goldman CME", len(pdf_goldman))
        with col_res2:
            st.metric("Santander CME", len(pdf_santander_cme))
        with col_res3:
            st.metric("Santander MEXDER", len(pdf_santander_mexder))
        with col_res4:
            st.metric("Scotia MEXDER", len(pdf_scotia))
        with col_res5:
            total_pdfs = len(pdf_goldman) + len(pdf_santander_cme) + len(pdf_santander_mexder) + len(pdf_scotia)
            st.metric("Total PDFs", total_pdfs)
        
        # Preview de PDFs si hay datos
        if total_pdfs > 0:
            st.markdown("### Preview de Datos Extraídos")
            
            todos_registros = []
            
            if pdf_goldman:
                recs = parse_goldman_cme(pdf_goldman)
                for r in recs:
                    todos_registros.append(r)
            
            if pdf_santander_cme:
                recs = parse_santander_cme(pdf_santander_cme)
                for r in recs:
                    todos_registros.append(r)
            
            if pdf_santander_mexder:
                recs = parse_santander_mexder(pdf_santander_mexder)
                for r in recs:
                    todos_registros.append(r)
            
            if pdf_scotia:
                recs = parse_scotia_mexder(pdf_scotia)
                for r in recs:
                    todos_registros.append(r)
            
            if todos_registros:
                df_preview = enriquecer(todos_registros, df_ref, fecha_valuacion)
                st.dataframe(
                    df_preview.head(20),
                    use_container_width=True,
                    hide_index=True
                )
                st.caption(f"Mostrando primeros 20 registros de {len(df_preview)} totales")
    
    # --- TAB 2: AIMS ---
    with tab2:
        st.markdown("### Carga de Archivo AIMS")
        file_aims = st.file_uploader("Archivo AIMS (Excel)", type=["xlsx", "xls"], key="aims_file")
        
        df_aims_ref = None
        if file_aims:
            try:
                df_aims_ref = leer_excel_cacheado(file_aims.getvalue())
                df_aims_ref.columns = [str(c).strip() for c in df_aims_ref.columns]
                df_aims_ref = df_aims_ref[
                    df_aims_ref["SOCIO"].map(normalizar_texto).isin(["SCOTIA", "SANTANDER"])
                ].copy()
                
                col_info1, col_info2 = st.columns(2)
                with col_info1:
                    st.info(f"Archivo cargado: {len(df_aims_ref)} filas (SCOTIA + SANTANDER)")
                with col_info2:
                    sieofres = sorted(df_aims_ref['SIEFORE'].unique())
                    st.info(f"SIEFOREs en AIMS: {', '.join(map(str, sieofres))}")
                
                st.markdown("### Preview AIMS")
                st.dataframe(df_aims_ref.head(10), use_container_width=True, hide_index=True)
            except Exception as e:
                st.error(f"Error al leer archivo AIMS: {e}")
                df_aims_ref = None
        else:
            st.info("No se ha cargado archivo AIMS. Las validaciones se realizarán sin referencia.")
    
    # --- TAB 3: IA COLATERAL ---
    with tab3:
        st.markdown("### Carga de Archivo IA Colateral")
        file_colateral = st.file_uploader("Archivo IA Colateral (Excel)", type=["xlsx", "xls"], key="colat_file")
        
        df_colat_ref = None
        if file_colateral:
            try:
                df_colat_ref = leer_excel_cacheado(file_colateral.getvalue())
                st.info(f"Archivo cargado: {len(df_colat_ref)} filas")
                st.markdown("### Preview IA Colateral")
                st.dataframe(df_colat_ref.head(10), use_container_width=True, hide_index=True)
            except Exception as e:
                st.error(f"Error al leer archivo IA Colateral: {e}")
                df_colat_ref = None
        else:
            st.info("No se ha cargado archivo IA Colateral. Las validaciones se realizarán sin referencia.")
    
    st.divider()
    
    # ========== ETAPA 2: VALIDACIÓN ==========
    st.markdown("## Etapa 2: Validación")
    
    col_btn1, col_btn2 = st.columns([1, 4])
    
    with col_btn1:
        btn_validar = st.button("Ejecutar Validación", type="primary", use_container_width=True)
    
    if btn_validar:
        # Procesar datos
        todos_registros = []
        
        if pdf_goldman:
            todos_registros += parse_goldman_cme(pdf_goldman)
        if pdf_santander_cme:
            todos_registros += parse_santander_cme(pdf_santander_cme)
        if pdf_santander_mexder:
            todos_registros += parse_santander_mexder(pdf_santander_mexder)
        if pdf_scotia:
            todos_registros += parse_scotia_mexder(pdf_scotia)
        
        if not todos_registros:
            st.warning("No se encontraron registros en los PDFs cargados.")
        else:
            # Enriquecer
            df_resultado = enriquecer(todos_registros, df_ref, fecha_valuacion)
            
            # Procesar AIMS
            pdf_scotia_dict = pdf_scotia if pdf_scotia else {}
            pdf_santander_mexder_dict = pdf_santander_mexder if pdf_santander_mexder else {}
            aims_pdf = procesar_aims_pdf(pdf_scotia_dict, pdf_santander_mexder_dict)
            df_aims_val = validar_aims(aims_pdf, df_aims_ref)
            
            # Procesar Colateral
            colateral_pdf = procesar_colateral_pdf(pdf_scotia_dict, pdf_santander_mexder_dict)
            df_colat_val = validar_colateral(colateral_pdf, df_colat_ref)
            
            # Guardar en session_state para generación de Excel
            st.session_state.df_resultado = df_resultado
            st.session_state.df_aims_val = df_aims_val
            st.session_state.df_colat_val = df_colat_val
            st.session_state.fecha_valuacion = fecha_valuacion
            st.session_state.pop("excel_bytes", None)
            st.session_state.pop("excel_nombre", None)
            
            st.success("Validación completada")
    
    if "df_resultado" in st.session_state:
        st.markdown("### Validación 1: Estados vs AIMS")
        if not st.session_state.df_aims_val.empty:
            st.dataframe(st.session_state.df_aims_val, use_container_width=True, hide_index=True)
        else:
            st.info("Sin datos de AIMS para validar.")
        
        st.markdown("### Validación 2: IA Colateral")
        if not st.session_state.df_colat_val.empty:
            st.dataframe(st.session_state.df_colat_val, use_container_width=True, hide_index=True)
        else:
            st.info("Sin datos de Colateral para validar.")
    
    st.divider()
    
    # ========== EXPORTACIÓN ==========
    st.markdown("## Exportación")
    
    if "df_resultado" in st.session_state:
        col_btn_export, col_info = st.columns([1, 3])
        
        with col_btn_export:
            btn_generar = st.button("Generar Excel", type="primary", use_container_width=True)
        
        if btn_generar:
            st.session_state.excel_bytes = generar_excel(
                st.session_state.df_resultado,
                st.session_state.df_aims_val,
                st.session_state.df_colat_val,
                st.session_state.fecha_valuacion
            )
            st.session_state.excel_nombre = (
                f"Validacion_Derivados_{st.session_state.fecha_valuacion.strftime('%d.%m.%Y')}.xlsx"
            )
        
        if "excel_bytes" in st.session_state:
            st.download_button(
                label="Descargar Excel",
                data=st.session_state.excel_bytes,
                file_name=st.session_state.excel_nombre,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            st.success(f"Archivo listo: {st.session_state.excel_nombre}")
    else:
        st.info("Ejecute la validación primero para generar el archivo Excel.")

if __name__ == "__main__":
    main()

